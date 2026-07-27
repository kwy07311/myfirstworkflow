"""
[개선판 - 수익률 관점 보강]

원본 스캐너는 "매수 신호를 찾는 기능"만 있고, 아래처럼 수익률에 직접 영향을 주는
요소들이 빠져 있었습니다. 이번 개선에서 추가/변경한 부분:

  1. [신규] 부실/저유동성 종목 사전 제외
     - 스팩(SPAC), 우선주, 거래대금이 너무 적은 종목은 슬리피지·변동성 리스크가 커서
       스캔 대상에서 미리 제외합니다. (MIN_LIQUIDITY_AMOUNT 로 조절 가능)

  2. [신규] ATR 기반 손절가/목표가 자동 계산
     - 신호가 뜬 모든 종목에 대해 ATR(14) 기준으로 손절가(진입가 - 1.5*ATR),
       목표가(진입가 + 3*ATR)를 계산해서 리스크/보상 비율(약 1:2)을 엑셀에 표시합니다.
     - "신호는 떴는데 손절선이 어디인지 몰라서 무한정 물타기"하는 상황을 방지하기 위함.

  3. [신규] 종목별 과거 신호 백테스트 (간이)
     - 오늘 신호가 뜬 종목에 한해서만, 해당 기법이 "이 종목의" 최근 1년 데이터에서
       과거에 몇 번 신호를 냈었고, 신호 이후 5거래일 뒤 평균 수익률/승률이 어땠는지
       계산해서 '과거승률(%)', '평균수익률(%)', '과거신호횟수' 컬럼으로 추가합니다.
     - 주의: 이건 "그 종목 자체의 과거 패턴 재현성"만 보는 매우 단순한 검증이며,
       전체 시장/전체 기간에 대한 정식 백테스트가 아닙니다. 표본이 적은 종목
       (과거신호횟수가 2~3회 이하)은 승률 수치를 신뢰하지 말고 참고만 하세요.
     - 전 종목(약 2000개) x 250일 백테스트를 다 돌리면 스캔이 매우 느려지므로,
       "오늘 이미 신호가 뜬 종목"에 한해서만 사후에 백테스트를 돌리는 방식으로
       속도 저하를 최소화했습니다.

  4. [신규] 시황(코스피 5일선 상회 여부) 정보를 전체 종목 공통 참고 컬럼으로 노출
     - 원본은 기법3(신정재 종가베팅)에만 시황 조건이 있었는데, 시장 레짐 필터는
       학계/실전 모두에서 가장 근거가 탄탄한 팩터 중 하나이므로 전체 결과에
       '코스피시황' 컬럼으로 노출합니다. (강제 제외는 하지 않음 - 종목별로 판단해
       하락장에서는 보수적으로 신호를 걸러 보는 용도)

  6. [신규] 기법별 전체시장 백테스트 (엑셀 2번째 시트)
     - "오늘 신호가 뜬 종목"만이 아니라 스캔 대상 전 종목에 대해, 9개 기법 각각을
       과거 MARKET_BACKTEST_LOOKBACK(기본 90거래일, 약 4~5개월)일 구간에서 재검증하고
       forward_days(기본 5거래일) 뒤 수익률을 계산해서, 기법별로 전체 시장 기준
       신호횟수/승률/평균수익률을 집계합니다 (엑셀 2번째 시트 "기법별_전체시장백테스트").
     - 네트워크 호출은 기존과 동일(종목당 1회)하고, 이미 받아온 히스토리 데이터로
       계산만 추가로 하는 것이라 스캔 속도에 미치는 영향을 최소화했습니다.
     - 단, 상장폐지/종목명 변경 등은 반영 못 하는 생존편향이 있고, 정식 시점별
       유니버스 재구성 백테스트는 아니라는 한계가 있습니다 (참고용).
     - 만약 이 요약에서 특정 기법의 전체시장 승률이 지속적으로 낮게 나온다면
       (예: 45% 미만, 평균수익률 마이너스) CONFIRMATION_TECHNIQUES 목록에서
       빼는 것을 고려해보세요 - 그 기법은 신호 개수만 늘릴 뿐 승률에 기여하지 않습니다.

  5. [주의사항 - 코드가 해결할 수 없는 부분]
     - 매도(청산) 타이밍, 포지션 사이징(종목당 투자 비중), 분산 투자(하루에 여러
       종목이 잡힐 때 특정 테마/섹터 쏠림 방지)는 이 스크립트의 범위를 벗어나는
       "운용 전략" 영역이라 자동화하지 않았습니다. 손절가/목표가는 참고용 가이드일
       뿐이며, 반드시 본인 리스크 허용 범위에 맞게 조정해서 사용하세요.
     - "매수신호" 개수가 많다고 승률이 비례해서 높아지는 것은 아닙니다. 여러 기법이
       비슷한 원리(양봉+거래량 급증)를 다르게 표현한 경우가 많아 서로 상관관계가
       높습니다. 개수보다는 과거승률/평균수익률 컬럼을 우선 참고하세요.

=====================================================================
[AI/사람 공통 안내] 새로운 매수 기법을 추가해달라는 요청을 받았다면?
=====================================================================
이 파일을 처음 보는 상태에서도 아래 순서만 따르면 매수 기법을 추가할 수 있다.
(예: "MACD 골든크로스 기법 추가해줘", "OO 영상에 나온 매매법 추가해줘" 같은 요청)

  1. has_xxx_signal(ohlcv_df, ...) -> bool  형태의 새 함수를 하나 작성한다.
     - 이 파일에서 CONFIRMATION_TECHNIQUES 바로 위에 정의된 함수들
       (has_candle_ma_breakout_signal, has_macd_buy_signal, has_bollinger_rsi_buy_signal 등)이
       전부 이 패턴을 따르고 있으니, 그대로 복사해서 참고하면 된다.
     - 첫 번째 인자 ohlcv_df는 'Open','High','Low','Close','Volume' 컬럼을 가진 pandas
       DataFrame이며, 인덱스는 날짜(과거->최근 순), 마지막 행(iloc[-1])이 가장 최근 거래일이다.
     - 데이터가 부족하면(예: 계산에 필요한 기간보다 짧으면) 조용히 False를 반환해야 한다.
       (에러를 던지면 안 됨 - check_stock()에서 try/except로 잡긴 하지만 스캔이 느려진다)
     - 지표 계산이 필요하면 calculate_macd(), calculate_bollinger_bands(), calculate_rsi()처럼
       "계산 함수(지표 컬럼을 추가한 DataFrame 반환) + 판단 함수(bool 반환)"로 분리하는 패턴을 따른다.
     - 함수 이름과 그 위 docstring에 어떤 매매법인지, 어디서 나온 조건인지(영상 제목 등) 간단히 적어둔다.
     - [중요] OHLCV만으로는 판단할 수 없는 정보(상장주식수, 거래대금 순위, 시장 전체 시황 등)가
       필요하다면, check_stock()/fast_find_eaten_candles()에서 스캔 시작 전에 "전체 시장 기준 1회"만
       계산해서 ohlcv_df에 부가 컬럼(예: 'Shares', 'IsLeading', 'MarketBullish')으로 실어서 넘긴다.
       (아래 has_extreme_volume_bb_breakout_signal / has_leading_stock_signal /
        has_jsj_closing_bet_signal 이 이 패턴의 예시다.) 이렇게 하면 함수 시그니처는
       여전히 ohlcv_df 하나만 받는 형태를 유지할 수 있고, 매 종목마다 추가 API를 호출하지 않아
       스캔 속도가 느려지지 않는다.

  2. 파일 하단의 CONFIRMATION_TECHNIQUES 리스트에 ("기법이름", 함수) 한 줄만 추가한다.
     기법이름은 엑셀의 '충족조건' 컬럼에 그대로 표시된다.

  3. 그 외 코드(check_stock, fast_find_eaten_candles, save_excel, main 등)는 절대 손댈 필요 없다.
     (단, 위 [중요] 항목처럼 부가 컬럼이 필요한 기법을 추가할 때는 예외적으로
      check_stock / fast_find_eaten_candles 에도 "전체 시장 1회 계산" 로직을 추가해야 한다.)
=====================================================================
"""

import os
import ssl
import smtplib
import logging
from email.message import EmailMessage
from datetime import datetime, timedelta
from concurrent.futures import ThreadPoolExecutor, as_completed

import pandas as pd
import FinanceDataReader as fdr
from openpyxl import Workbook
from openpyxl.styles import Font
from urllib.parse import quote

# =========================================================
# ▼▼▼ 비밀번호는 GitHub Secrets에서 환경변수로 주입됩니다 ▼▼▼
# =========================================================
NAVER_EMAIL = os.environ.get("NAVER_EMAIL", "")
NAVER_PASSWORD = os.environ.get("NAVER_PASSWORD", "")
RECEIVER_EMAIL = os.environ.get("RECEIVER_EMAIL", "")
# =========================================================

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
LOG_PATH = os.path.join(SCRIPT_DIR, "stock_auto_mail.log")

logging.basicConfig(
    filename=LOG_PATH,
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    encoding="utf-8-sig",
)

# =========================================================
# 전체 시장 기준 1회 계산 파라미터 (기법2/기법3에서 사용)
# =========================================================
LEADING_STOCK_TOP_PCT = 0.03      # 거래대금 상위 3%를 '주도주 후보군'으로 정의

# =========================================================
# [신규] 리스크 관리 / 종목 필터링 파라미터
# =========================================================
MIN_LIQUIDITY_AMOUNT = 500_000_000   # 최소 거래대금(원). 이 미만 종목은 슬리피지 위험으로 제외
ATR_PERIOD = 14
ATR_STOP_MULT = 1.5                  # 손절 = 진입가 - ATR_STOP_MULT * ATR
ATR_TARGET_MULT = 3.0                # 목표 = 진입가 + ATR_TARGET_MULT * ATR (약 1:2 손익비)
BACKTEST_FORWARD_DAYS = 5            # 신호 이후 며칠 뒤 수익률로 검증할지
MARKET_BACKTEST_LOOKBACK = 90        # [신규] 전체 시장 공통 백테스트 기간 (모든 스캔 종목에 적용, 속도 고려해 90일로 제한)
                                      # 종목별 화면 표시용 승률/평균수익률도 이 값을 그대로 재사용한다


# =========================================================
# 매매기법 정의 구역
# =========================================================

def calculate_double_bollinger_bands(df, wb_period=4, wb_std=4, bb_period=20, bb_std=2):
    """
    더블 볼린저 밴드 지표 계산.
    - WB(4,4): 시가(Open) 기준, 단기 극한 변곡·돌파 포착용
    - 기본 BB(20,2): 종가(Close) 기준, 표준 레인지/평균회귀 기준
    """
    out = df.copy()

    wb_source = out['Open'].astype(float)
    wb_ma = wb_source.rolling(window=wb_period).mean()
    wb_stdval = wb_source.rolling(window=wb_period).std()
    out['WB_Mid'] = wb_ma
    out['WB_Upper'] = wb_ma + wb_std * wb_stdval
    out['WB_Lower'] = wb_ma - wb_std * wb_stdval

    bb_source = out['Close'].astype(float)
    bb_ma = bb_source.rolling(window=bb_period).mean()
    bb_stdval = bb_source.rolling(window=bb_period).std()
    out['BB_Mid'] = bb_ma
    out['BB_Upper'] = bb_ma + bb_std * bb_stdval
    out['BB_Lower'] = bb_ma - bb_std * bb_stdval

    return out


def has_double_bollinger_buy_signal(
    ohlcv_df,
    wb_period=4, wb_std=4, bb_period=20, bb_std=2,
    resistance_lookback=20, wick_ratio_threshold=1.5, support_tolerance_pct=1.0,
) -> bool:
    """
    [더블 볼린저 밴드(WB) 매매법] - 유튜브 '김직선 - 나스닥 트레이더' 채널
    https://www.youtube.com/watch?v=t800Joz9GHw

    WB(4,4,Open,빨간선) + 기본 볼린저(20,2,Close,흰선)를 함께 써서
    매수 관점에서 아래 두 패턴 중 하나라도 만족하면 True.

    [패턴 A] 하단 변곡(반전) 매수
      - 오늘 저가가 WB 하단 또는 BB 하단을 터치/이탈
      - 종가는 두 밴드 안쪽으로 복귀 (아래꼬리 긴 반전 마감)
      - 직전 저가(매물대) 대비 유의미한 신저가를 만들지 않음 (돌파 실패 확인)

    [패턴 B] 상단 진짜 돌파 후 추세 매수
      - 몸통(시가->종가)이 WB 상단과 BB 상단을 모두 꽉 채우며 강하게 돌파
      - 직전 고점(매물대, resistance_lookback일 종가 최고가)도 종가 기준 함께 돌파

    데이터 부족 시 조용히 False 반환.
    """
    min_len = max(wb_period, bb_period, resistance_lookback) + 2
    if ohlcv_df is None or len(ohlcv_df) < min_len:
        return False

    df = calculate_double_bollinger_bands(ohlcv_df, wb_period, wb_std, bb_period, bb_std)

    today = df.iloc[-1]

    if pd.isna(today['WB_Lower']) or pd.isna(today['BB_Lower']):
        return False

    # ---------- 패턴 A: 하단 변곡(반전) 매수 ----------
    touched_lower_band = (today['Low'] <= today['WB_Lower']) or (today['Low'] <= today['BB_Lower'])
    closed_back_inside = (today['Close'] > today['WB_Lower']) and (today['Close'] > today['BB_Lower'])

    body_high = max(today['Open'], today['Close'])
    body_low = min(today['Open'], today['Close'])
    body_size = body_high - body_low
    lower_wick = body_low - today['Low']
    has_long_lower_wick = body_size > 0 and (lower_wick > body_size * wick_ratio_threshold)

    prior_low = df['Low'].iloc[-(resistance_lookback + 1):-1].min()
    support_not_broken = today['Low'] >= prior_low * (1 - support_tolerance_pct / 100)

    reversal_pattern = bool(
        touched_lower_band and closed_back_inside and has_long_lower_wick and support_not_broken
    )

    # ---------- 패턴 B: 상단 진짜 돌파 후 추세 매수 ----------
    body_breaks_both_bands = (
        (today['Open'] <= today['WB_Upper']) and (today['Close'] > today['WB_Upper'])
        and (today['Open'] <= today['BB_Upper']) and (today['Close'] > today['BB_Upper'])
    )

    prior_high_close = df['Close'].iloc[-(resistance_lookback + 1):-1].max()
    breaks_resistance = today['Close'] > prior_high_close

    breakout_pattern = bool(body_breaks_both_bands and breaks_resistance)

    return reversal_pattern or breakout_pattern


def has_candle_ma_breakout_signal(ohlcv_df, volume_multiplier=1.5) -> bool:
    """
    [기존 1차 조건] 오늘 양봉이면서 시가는 MA5/MA20 아래, 종가는 MA5/MA20 위로 돌파,
    거래량은 전일 대비 volume_multiplier배 이상 증가.
    """
    if ohlcv_df is None or len(ohlcv_df) < 21:
        return False

    df = ohlcv_df.copy()
    df['MA5'] = df['Close'].rolling(window=5).mean()
    df['MA20'] = df['Close'].rolling(window=20).mean()

    today = df.iloc[-1]
    yesterday = df.iloc[-2]

    is_bullish = today['Close'] > today['Open']
    breaks_ma5 = (today['Open'] < today['MA5']) and (today['Close'] > today['MA5'])
    breaks_ma20 = (today['Open'] < today['MA20']) and (today['Close'] > today['MA20'])
    volume_up = today['Volume'] > (yesterday['Volume'] * volume_multiplier)

    return bool(is_bullish and breaks_ma5 and breaks_ma20 and volume_up)


def has_pullback_after_breakout_signal(
    ohlcv_df,
    ma5_period=5, ma20_period=20, ma60_period=60,
    lookback_days=15, volume_surge_mult=2.0, body_surge_pct=5.0,
    body_damage_tolerance=0.3, ma5_support_tolerance=0.02,
) -> bool:
    """
    [정배열 + 대량거래량 장대양봉 눌림목 기법]
    """
    min_len = ma60_period + lookback_days
    if ohlcv_df is None or len(ohlcv_df) < min_len:
        return False

    df = ohlcv_df.copy()
    df['MA5'] = df['Close'].rolling(window=ma5_period).mean()
    df['MA20'] = df['Close'].rolling(window=ma20_period).mean()
    df['MA60'] = df['Close'].rolling(window=ma60_period).mean()
    df['AvgVolume20'] = df['Volume'].rolling(window=20).mean()
    df['PrevHigh20'] = df['Close'].shift(1).rolling(window=20).max()

    today = df.iloc[-1]
    if pd.isna(today['MA60']):
        return False

    is_uptrend_alignment = (today['MA5'] > today['MA20']) and (today['MA20'] > today['MA60'])
    if not is_uptrend_alignment:
        return False

    recent = df.iloc[-(lookback_days + 1):-1]
    is_bullish = recent['Close'] > recent['Open']
    body_pct = (recent['Close'] - recent['Open']) / recent['Open'] * 100
    volume_surge = recent['Volume'] > (recent['AvgVolume20'] * volume_surge_mult)
    breaks_resistance = recent['Close'] > recent['PrevHigh20']

    breakout_candidates = recent[is_bullish & (body_pct >= body_surge_pct) & volume_surge & breaks_resistance]
    if breakout_candidates.empty:
        return False

    breakout_candle = breakout_candidates.iloc[-1]
    body_low = min(breakout_candle['Open'], breakout_candle['Close'])
    body_high = max(breakout_candle['Open'], breakout_candle['Close'])
    body_range = body_high - body_low
    if body_range <= 0:
        return False

    min_allowed_close = body_low - body_range * body_damage_tolerance
    body_not_damaged = today['Close'] >= min_allowed_close
    near_ma5_support = today['Close'] >= today['MA5'] * (1 - ma5_support_tolerance)
    volume_calmed = today['Volume'] < breakout_candle['Volume']

    return bool(body_not_damaged and near_ma5_support and volume_calmed)


def calculate_macd(df, close_col='Close', fast=12, slow=26, signal=9):
    out = df.copy()
    close = out[close_col].astype(float)
    ema_fast = close.ewm(span=fast, adjust=False).mean()
    ema_slow = close.ewm(span=slow, adjust=False).mean()
    out['MACD'] = ema_fast - ema_slow
    out['Signal'] = out['MACD'].ewm(span=signal, adjust=False).mean()
    out['Histogram'] = out['MACD'] - out['Signal']
    return out


def has_macd_buy_signal(ohlcv_df) -> bool:
    """
    [MACD 기법] 골든크로스 또는 제로라인 상향 돌파
    """
    if ohlcv_df is None or len(ohlcv_df) < 2:
        return False

    df_macd = calculate_macd(ohlcv_df)
    today = df_macd.iloc[-1]
    yesterday = df_macd.iloc[-2]

    golden_cross = (yesterday['MACD'] < yesterday['Signal']) and (today['MACD'] > today['Signal'])
    zero_line_breakout = (yesterday['MACD'] < 0) and (today['MACD'] > 0)

    return bool(golden_cross or zero_line_breakout)


def calculate_bollinger_bands(df, close_col='Close', period=20, num_std=2):
    out = df.copy()
    close = out[close_col].astype(float)
    ma = close.rolling(window=period).mean()
    std = close.rolling(window=period).std()
    out['BB_Mid'] = ma
    out['BB_Upper'] = ma + num_std * std
    out['BB_Lower'] = ma - num_std * std
    return out


def calculate_rsi(df, close_col='Close', period=14):
    out = df.copy()
    delta = out[close_col].astype(float).diff()
    gain = delta.clip(lower=0).rolling(window=period).mean()
    loss = (-delta.clip(upper=0)).rolling(window=period).mean()
    rs = gain / loss
    out['RSI'] = 100 - (100 / (1 + rs))
    return out


def has_bollinger_rsi_buy_signal(ohlcv_df, bb_period=20, bb_std=2, rsi_period=14, rsi_buy_threshold=25) -> bool:
    """
    [볼린저밴드 + RSI 기법]
    """
    min_len = max(bb_period, rsi_period) + 1
    if ohlcv_df is None or len(ohlcv_df) < min_len:
        return False

    df_bb = calculate_bollinger_bands(ohlcv_df, period=bb_period, num_std=bb_std)
    df_rsi = calculate_rsi(ohlcv_df, period=rsi_period)

    today_close = df_bb['Close'].iloc[-1]
    today_lower = df_bb['BB_Lower'].iloc[-1]
    today_rsi = df_rsi['RSI'].iloc[-1]

    if pd.isna(today_lower) or pd.isna(today_rsi):
        return False

    price_below_lower_band = today_close < today_lower
    rsi_oversold = today_rsi <= rsi_buy_threshold

    return bool(price_below_lower_band and rsi_oversold)


def has_extreme_volume_bb_breakout_signal(
    ohlcv_df,
    bb_period=20, bb_std=1,
    volume_surge_mult=7.0, year_high_lookback=240,
    wick_body_ratio_max=1.0, turnover_min_pct=5.0,
) -> bool:
    """
    [매매기법1] 역대급 거래량 + 볼린저(20,1) 상단 강돌파

    - 역대급 거래량: 최근 year_high_lookback(기본 240봉=1년) 중 최고 거래량이거나,
      20일 평균 거래량 대비 volume_surge_mult(기본 700%) 이상 급증.
    - 볼린저 밴드(20,1) 상단선을 강하게 돌파하는 양봉.
    - 캔들 모양: 위꼬리가 몸통보다 길지 않음(짧거나 비등).
    - 회전율(거래량/상장주식수)이 높은 종목 우대.
      * 회전율 계산에는 상장주식수가 필요하므로, check_stock()에서 ohlcv_df에
        'Shares'(상장주식수) 컬럼을 미리 실어서 넘긴다. 'Shares' 컬럼이 없으면
        (데이터 미제공 종목 등) 회전율 조건은 건너뛰고 나머지 조건만으로 판단한다.

    데이터 부족 시 조용히 False 반환.
    """
    min_len = max(bb_period, 21) + 1
    if ohlcv_df is None or len(ohlcv_df) < min_len:
        return False

    df = calculate_bollinger_bands(ohlcv_df, period=bb_period, num_std=bb_std)
    today = df.iloc[-1]

    if pd.isna(today['BB_Upper']):
        return False

    is_bullish = today['Close'] > today['Open']

    # ---------- 역대급 거래량 ----------
    vol_window = df['Volume'].iloc[-min(year_high_lookback, len(df)):]
    is_year_high_volume = bool(today['Volume'] >= vol_window.max())

    avg20_volume = df['Volume'].iloc[-21:-1].mean()
    is_volume_surge = bool(today['Volume'] >= avg20_volume * volume_surge_mult)

    extreme_volume = is_year_high_volume or is_volume_surge

    # ---------- 볼린저(20,1) 상단 강돌파 ----------
    strong_bb_breakout = bool(today['Close'] > today['BB_Upper'])

    # ---------- 캔들 모양 (위꼬리 <= 몸통) ----------
    body_size = today['Close'] - today['Open']
    upper_wick = today['High'] - max(today['Open'], today['Close'])
    clean_candle = bool(body_size > 0 and upper_wick <= body_size * wick_body_ratio_max)

    # ---------- 회전율 (상장주식수 있을 때만 체크) ----------
    turnover_ok = True
    if 'Shares' in ohlcv_df.columns:
        shares = ohlcv_df['Shares'].iloc[-1]
        if pd.notna(shares) and shares > 0:
            turnover_rate = today['Volume'] / shares * 100
            turnover_ok = bool(turnover_rate >= turnover_min_pct)

    return bool(is_bullish and extreme_volume and strong_bb_breakout and clean_candle and turnover_ok)


def has_leading_stock_signal(
    ohlcv_df,
    benchmark_lookback=20, ma5_period=5, ma10_period=10,
    support_tolerance_pct=2.0, close_strength_pct=70.0,
    require_leading=True,
) -> bool:
    """
    [매매기법2] 주도주(거래대금 상위) 기준봉 돌파 / 눌림목 지지

    원본 기법의 "주도 테마·재료" 여부는 뉴스/공시 데이터가 없어 자동 판단이 불가능하므로
    제외했고, 대신 '당일 거래대금 상위 종목군에 속하는지'로 대체했다.
    (상위 그룹 판정은 fast_find_eaten_candles()에서 전체 시장을 대상으로 1회만 계산해
     ohlcv_df에 'IsLeading' 컬럼으로 실어서 넘긴다.)

    아래 두 패턴 중 하나라도 만족하면 True.
    [패턴 A] 최근 benchmark_lookback일 내 거래량이 가장 컸던 '기준봉'의 고점을
             오늘 종가가 강하게(고가권 마감) 돌파.
    [패턴 B] 5일/10일 이동평균선 부근에서 지지받으며 고가권으로 마감(눌림목).

    데이터 부족 또는 'IsLeading'이 아닌 종목은(require_leading=True일 때) False.
    """
    min_len = max(benchmark_lookback, ma10_period) + 2
    if ohlcv_df is None or len(ohlcv_df) < min_len:
        return False

    if require_leading:
        if 'IsLeading' not in ohlcv_df.columns or not bool(ohlcv_df['IsLeading'].iloc[-1]):
            return False

    df = ohlcv_df.copy()
    df['MA5'] = df['Close'].rolling(window=ma5_period).mean()
    df['MA10'] = df['Close'].rolling(window=ma10_period).mean()

    today = df.iloc[-1]

    day_range = today['High'] - today['Low']
    close_position_pct = ((today['Close'] - today['Low']) / day_range * 100) if day_range > 0 else 100.0
    strong_close = bool(close_position_pct >= close_strength_pct)

    # ---------- 패턴 A: 기준봉 고점 돌파 ----------
    recent = df.iloc[-(benchmark_lookback + 1):-1]
    pattern_a = False
    if not recent.empty and recent['Volume'].notna().any():
        benchmark_idx = recent['Volume'].idxmax()
        benchmark_high = recent.loc[benchmark_idx, 'High']
        pattern_a = bool(today['Close'] > benchmark_high and today['Close'] > today['Open'] and strong_close)

    # ---------- 패턴 B: 이동평균선 눌림목 지지 ----------
    pattern_b = False
    if pd.notna(today['MA5']) and pd.notna(today['MA10']):
        near_ma_support = (
            today['Close'] >= today['MA5'] * (1 - support_tolerance_pct / 100)
        ) or (
            today['Close'] >= today['MA10'] * (1 - support_tolerance_pct / 100)
        )
        pattern_b = bool(near_ma_support and strong_close)

    return pattern_a or pattern_b


def has_jsj_closing_bet_signal(
    ohlcv_df,
    new_high_lookback=120, gap_tolerance_pct=3.0,
    upper_wick_max_pct=15.0, volume_surge_mult=3.0,
    consolidation_period=60, consolidation_range_max_pct=20.0,
    require_market_bullish=True,
) -> bool:
    """
    [매매기법3] 신정재 종가베팅 (기술적 조건만 구현)

    원문의 6가지 조건 중 아래 4가지는 OHLCV로 판단 가능해 구현했다.
      1) 신고가/전고점 돌파
      2) 전고점과의 이격거리가 좁음(멀리 갭 뜨지 않은 돌파)
      3) 위꼬리가 거의 없는 깔끔한 양봉
      4) 평소 대비 거래량 급증
      5) 충분한 기간조정(consolidation_period일 레인지가 좁았는지) 후 상승

    "재료(뉴스·테마)"는 자동 판단이 불가능해 제외했다.
    "시황(지수 5일선)"은 스캔 시작 전 전체 시장 기준 1회만 계산해 'MarketBullish'
    컬럼으로 실어서 넘기며, require_market_bullish=True일 때만 조건에 반영한다.
    "수급(외국인/기관 순매수)"은 이번 구현에서는 제외했다 (안정성 우선, pykrx 등
    별도 의존성 필요 - 추후 필요하면 이 함수에 조건을 추가하면 된다).

    데이터 부족 시 조용히 False 반환.
    """
    min_len = max(new_high_lookback, consolidation_period) + 5
    if ohlcv_df is None or len(ohlcv_df) < min_len:
        return False

    df = ohlcv_df
    today = df.iloc[-1]

    # ---------- 신고가 + 이격거리 ----------
    prior = df.iloc[-(new_high_lookback + 1):-1]
    prior_high = prior['Close'].max()
    if pd.isna(prior_high) or prior_high <= 0:
        return False

    is_new_high = bool(today['Close'] > prior_high)
    if not is_new_high:
        return False

    gap_pct = (today['Close'] - prior_high) / prior_high * 100
    narrow_gap = bool(gap_pct <= gap_tolerance_pct)

    # ---------- 깔끔한 양봉 ----------
    body_size = today['Close'] - today['Open']
    upper_wick = today['High'] - max(today['Open'], today['Close'])
    clean_candle = bool(body_size > 0 and upper_wick <= body_size * (upper_wick_max_pct / 100))

    # ---------- 거래량 급증 ----------
    avg20_volume = df['Volume'].iloc[-21:-1].mean()
    volume_surge = bool(today['Volume'] >= avg20_volume * volume_surge_mult)

    # ---------- 기간조정 (돌파 직전 consolidation_period일 레인지) ----------
    consolidation_window = df.iloc[-(consolidation_period + 1):-1]
    cons_high = consolidation_window['Close'].max()
    cons_low = consolidation_window['Close'].min()
    had_consolidation = False
    if pd.notna(cons_high) and pd.notna(cons_low) and cons_low > 0:
        cons_range_pct = (cons_high - cons_low) / cons_low * 100
        had_consolidation = bool(cons_range_pct <= consolidation_range_max_pct)

    # ---------- 시황 (지수, 있을 때만 체크) ----------
    market_ok = True
    if require_market_bullish and 'MarketBullish' in ohlcv_df.columns:
        market_ok = bool(ohlcv_df['MarketBullish'].iloc[-1])

    return bool(narrow_gap and clean_candle and volume_surge and had_consolidation and market_ok)


def calculate_swing_points(df, window=3):
    """
    프랙탈(fractal) 방식으로 스윙 고점/저점을 계산한다.
    i번째 봉의 High가 좌우 window개 봉을 통틀어 유일한 최댓값이면 스윙 고점,
    Low가 유일한 최솟값이면 스윙 저점으로 표시한다.
    (마지막 window개 봉은 좌우 확인이 불가능해 스윙 여부를 판단하지 않는다 -
     과거 데이터 전체를 놓고 계산하므로 미래 데이터를 미리 보는 문제는 없다.)
    """
    out = df.copy()
    highs = out['High'].values
    lows = out['Low'].values
    n = len(out)
    swing_high = [False] * n
    swing_low = [False] * n

    for i in range(window, n - window):
        seg_h = highs[i - window:i + window + 1]
        if highs[i] == seg_h.max() and (seg_h == highs[i]).sum() == 1:
            swing_high[i] = True
        seg_l = lows[i - window:i + window + 1]
        if lows[i] == seg_l.min() and (seg_l == lows[i]).sum() == 1:
            swing_low[i] = True

    out['SwingHigh'] = swing_high
    out['SwingLow'] = swing_low
    return out


def get_market_structure(df_with_swings, lookback=120, min_swings=2):
    """
    최근 lookback일(오늘 제외) 안의 확정 스윙 고점/저점 중 마지막 min_swings개를 비교해서
    'up'(HH+HL), 'down'(LH+LL), 'sideways', 또는 스윙이 부족하면 'unknown'을 반환한다.
    """
    recent = df_with_swings.iloc[-(lookback + 1):-1]
    highs = recent[recent['SwingHigh']]['High']
    lows = recent[recent['SwingLow']]['Low']

    if len(highs) < min_swings or len(lows) < min_swings:
        return 'unknown'

    last_highs = highs.iloc[-min_swings:]
    last_lows = lows.iloc[-min_swings:]

    hh = all(last_highs.iloc[i] < last_highs.iloc[i + 1] for i in range(len(last_highs) - 1))
    hl = all(last_lows.iloc[i] < last_lows.iloc[i + 1] for i in range(len(last_lows) - 1))
    lh = all(last_highs.iloc[i] > last_highs.iloc[i + 1] for i in range(len(last_highs) - 1))
    ll = all(last_lows.iloc[i] > last_lows.iloc[i + 1] for i in range(len(last_lows) - 1))

    if hh and hl:
        return 'up'
    if lh and ll:
        return 'down'
    return 'sideways'


def is_near_round_number(price, tolerance_pct=0.5) -> bool:
    """
    가격이 심리적으로 딱 떨어지는 '라운드 피겨' 근처인지 확인한다.
    가격대에 따라 단위를 다르게 잡는다(10만원 이상->1만원 단위, 1만원 이상->1천원 단위 ...).
    """
    if price is None or price <= 0:
        return False
    if price >= 100000:
        step = 10000
    elif price >= 10000:
        step = 1000
    elif price >= 1000:
        step = 100
    elif price >= 100:
        step = 10
    else:
        step = 1

    nearest = round(price / step) * step
    if nearest == 0:
        return False
    return bool(abs(price - nearest) / price * 100 <= tolerance_pct)


def has_price_action_pullback_signal(
    ohlcv_df,
    swing_window=3, structure_lookback=120, min_swings=2,
    support_lookback=60, zone_tolerance_pct=1.5,
    round_number_tolerance_pct=0.5,
    pinbar_wick_ratio=2.0, volume_surge_mult=1.2,
    require_uptrend_structure=True,
) -> bool:
    """
    [프라이스 액션(Price Action) 눌림목 / SR Flip 매수 기법]
    https://www.youtube.com/watch?v=S4DYqE9Q5Zc

    다음을 종합해서, "상승추세에서 지지구간(또는 SR Flip / 라운드 피겨)까지 눌린 뒤
    반등 캔들(핀바 또는 장악형) + 거래량 증가로 확인되는" 매수 시점을 판단한다.
      1) 시장 구조(HH/HL 상승추세 - 스윙 고점/저점 기반, 명확한 하락추세면 제외)
      2) 지지 구간: 최근 support_lookback일 내 스윙 저점 중 오늘 저가와 가장 가까운 값과
         zone_tolerance_pct 이내로 근접
      3) SR Flip: 과거 스윙 고점(저항) 중 이후 확실히 돌파(종가 기준 +2% 이상)된 레벨을
         오늘 저가가 다시 눌러줌(저항->지지 전환 재테스트)
      4) 라운드 피겨: 오늘 저가가 심리적 라운드 넘버 근처
      5) 캔들 확인: 불리시 핀바(아래꼬리가 몸통의 pinbar_wick_ratio배 이상, 위꼬리는 짧음)
         또는 불리시 장악형(오늘 양봉이 전일 음봉 몸통을 완전히 감쌈)
      6) 거래량 동반: 오늘 거래량이 20일 평균 대비 volume_surge_mult배 이상

    [주의 - 근사치임을 밝힘]
      - '유동성/SFP'는 별도로 정밀하게 구현하지 않았고, "종가가 지지선 위로 복귀"하는
        조건(위 2/3/4번 + 캔들 확인)으로만 근사했다. 분봉/호가 단위의 진짜 유동성 흡수
        여부는 일봉 데이터로는 판단할 수 없다.
      - 지지/저항 구간은 스윙 고점/저점(프랙탈) 기준 근사치이며, 사람이 차트를 보고
        직접 긋는 정교한 구간과는 다를 수 있다.
      - 멀티 타임프레임 분석은 하지 않는다(일봉만 사용).

    데이터 부족 시 조용히 False 반환.
    """
    min_len = max(structure_lookback, support_lookback) + swing_window * 2 + 5
    if ohlcv_df is None or len(ohlcv_df) < min_len:
        return False

    df = calculate_swing_points(ohlcv_df, window=swing_window)
    today = df.iloc[-1]
    yesterday = df.iloc[-2]

    # ---------- 시장 구조 ----------
    structure = get_market_structure(df, lookback=structure_lookback, min_swings=min_swings)
    if require_uptrend_structure and structure == 'down':
        return False

    # ---------- 지지 구간 / SR Flip / 라운드 피겨 ----------
    recent_for_support = df.iloc[-(support_lookback + 1):-1]

    swing_lows = recent_for_support[recent_for_support['SwingLow']]['Low']
    near_support_zone = False
    if not swing_lows.empty:
        nearest_low = swing_lows.loc[(swing_lows - today['Low']).abs().idxmin()]
        near_support_zone = bool(abs(today['Low'] - nearest_low) / nearest_low * 100 <= zone_tolerance_pct)

    swing_highs = recent_for_support[recent_for_support['SwingHigh']]['High']
    sr_flip_zone = False
    for level in swing_highs:
        was_broken = bool((recent_for_support['Close'] > level * 1.02).any())
        now_retesting = bool(abs(today['Low'] - level) / level * 100 <= zone_tolerance_pct)
        if was_broken and now_retesting:
            sr_flip_zone = True
            break

    near_round = is_near_round_number(today['Low'], tolerance_pct=round_number_tolerance_pct)

    valid_zone = near_support_zone or sr_flip_zone or near_round
    if not valid_zone:
        return False

    # ---------- 캔들 확인 (핀바 또는 장악형) ----------
    body_size = abs(today['Close'] - today['Open'])
    lower_wick = min(today['Open'], today['Close']) - today['Low']
    upper_wick = today['High'] - max(today['Open'], today['Close'])

    is_bullish_pinbar = bool(
        lower_wick > 0
        and lower_wick >= max(body_size, 1e-9) * pinbar_wick_ratio
        and upper_wick <= lower_wick * 0.5
        and today['Close'] >= today['Open']
    )
    is_bullish_engulfing = bool(
        yesterday['Close'] < yesterday['Open']
        and today['Close'] > today['Open']
        and today['Open'] <= yesterday['Close']
        and today['Close'] >= yesterday['Open']
    )
    confirmation_candle = is_bullish_pinbar or is_bullish_engulfing
    if not confirmation_candle:
        return False

    # ---------- 거래량 동반 ----------
    avg20_volume = df['Volume'].iloc[-21:-1].mean()
    volume_confirmed = bool(today['Volume'] >= avg20_volume * volume_surge_mult)

    return bool(confirmation_candle and volume_confirmed)


CONFIRMATION_TECHNIQUES = [
    ("양봉+MA돌파+거래량", has_candle_ma_breakout_signal),
    ("MACD", has_macd_buy_signal),
    ("볼린저밴드+RSI", has_bollinger_rsi_buy_signal),
    ("정배열+장대양봉눌림목", has_pullback_after_breakout_signal),
    ("더블볼린저밴드(WB)", has_double_bollinger_buy_signal),
    ("역대급거래량+볼린저상단강돌파", has_extreme_volume_bb_breakout_signal),  # 매매기법1
    ("주도주+거래대금상위", has_leading_stock_signal),                        # 매매기법2
    ("신정재종가베팅", has_jsj_closing_bet_signal),                           # 매매기법3
    ("프라이스액션눌림목+SRFlip", has_price_action_pullback_signal),          # 프라이스 액션 기법
]


# =========================================================
# [신규] 리스크 관리 (ATR 손절/목표가) & 종목별 과거 신호 백테스트
# =========================================================

def calculate_atr(df, period=ATR_PERIOD):
    """
    ATR(Average True Range) 계산. 변동성 기반 손절/목표가 산정에 사용.
    """
    out = df.copy()
    prev_close = out['Close'].shift(1)
    tr = pd.concat([
        out['High'] - out['Low'],
        (out['High'] - prev_close).abs(),
        (out['Low'] - prev_close).abs(),
    ], axis=1).max(axis=1)
    out['ATR'] = tr.rolling(window=period).mean()
    return out


def get_risk_levels(ohlcv_df):
    """
    오늘 종가 기준 ATR 손절가/목표가/손익비를 계산한다.
    데이터 부족 시 (None, None, None) 반환.
    """
    if ohlcv_df is None or len(ohlcv_df) < ATR_PERIOD + 2:
        return None, None, None

    df_atr = calculate_atr(ohlcv_df)
    today_close = df_atr['Close'].iloc[-1]
    today_atr = df_atr['ATR'].iloc[-1]

    if pd.isna(today_atr) or today_atr <= 0:
        return None, None, None

    stop_loss = today_close - ATR_STOP_MULT * today_atr
    target = today_close + ATR_TARGET_MULT * today_atr
    risk = today_close - stop_loss
    reward = target - today_close
    rr_ratio = round(reward / risk, 2) if risk > 0 else None

    return int(round(stop_loss)), int(round(target)), rr_ratio


def backtest_technique_raw(full_df, check_fn, forward_days=BACKTEST_FORWARD_DAYS,
                           max_lookback=MARKET_BACKTEST_LOOKBACK):
    """
    [핵심 백테스트 함수] 이 종목의 과거 데이터에서 check_fn이 신호를 낸 시점들을 찾아,
    각 신호 발생 이후 forward_days 거래일 뒤 수익률을 계산한다.

    이 함수는 "종목 하나"에 대한 결과를 raw(승수, 표본수, 수익률합계)로 반환한다.
    - 종목별로 그대로 보여주는 용도로도 쓸 수 있고,
    - 여러 종목의 raw 결과를 그대로 더하면(wins 합, total 합, sum_returns 합)
      "기법별 전체 시장 통합 승률/평균수익률"을 정확하게 집계할 수 있다
      (비율을 미리 반올림해서 평균내는 방식보다 통계적으로 정확함).

    주의: 상장폐지 종목 등은 반영되지 않는 생존편향이 있고, 정식 시점별 유니버스
    재구성 백테스트는 아니다 (참고용 근사치).

    반환: (wins:int, total:int, sum_returns:float) — 표본 없으면 (0, 0, 0.0)
    """
    n = len(full_df)
    start_idx = max(30, n - max_lookback)
    end_idx = n - forward_days  # 미래 데이터가 있어야 수익률 계산 가능
    if end_idx <= start_idx:
        return 0, 0, 0.0

    wins = 0
    total = 0
    sum_returns = 0.0
    for i in range(start_idx, end_idx):
        window = full_df.iloc[:i + 1]
        try:
            if check_fn(window):
                entry = full_df['Close'].iloc[i]
                exit_price = full_df['Close'].iloc[i + forward_days]
                if entry > 0:
                    ret = (exit_price - entry) / entry * 100
                    sum_returns += ret
                    total += 1
                    if ret > 0:
                        wins += 1
        except Exception:
            continue

    return wins, total, sum_returns


def summarize_raw(wins, total, sum_returns):
    """(wins, total, sum_returns) -> (승률%, 평균수익률%) / 표본 없으면 (None, None)"""
    if total == 0:
        return None, None
    return round(wins / total * 100, 1), round(sum_returns / total, 2)


def get_market_bullish_flag(end_date: str) -> bool:
    """
    코스피 지수가 5일 이동평균선 위에 있는지(시황 상승/반등 여부)를
    스캔 시작 전 딱 1번만 조회해서 판단한다. 조회 실패 시 안전하게 True(조건 미적용)로
    처리해서 전체 스캔이 실패하지 않도록 한다.
    """
    try:
        start = (datetime.strptime(end_date, '%Y-%m-%d') - timedelta(days=30)).strftime('%Y-%m-%d')
        idx_df = fdr.DataReader('KS11', start=start, end=end_date)
        idx_df['MA5'] = idx_df['Close'].rolling(window=5).mean()
        last = idx_df.iloc[-1]
        if pd.isna(last['MA5']):
            return True
        return bool(last['Close'] >= last['MA5'])
    except Exception as e:
        logging.warning(f"코스피 지수 조회 실패, 시황 조건 기본값(True)으로 진행: {e}")
        return True


def check_stock(row, start_date, end_date, market_bullish=True):
    """
    반환값: (result_dict_or_None, tech_stats_list)
      - result_dict_or_None: 오늘 신호가 뜬 경우에만 dict, 아니면 None (엑셀 메인 시트용)
      - tech_stats_list: 이 종목의 9개 기법 전체에 대한 (기법명, wins, total, sum_returns) 리스트.
        오늘 신호 매칭 여부와 무관하게 "전체 시장 기준 기법별 백테스트" 집계를 위해
        스캔한 모든 종목에서 항상 계산한다. (네트워크 호출은 추가되지 않고, 이미 받아온
        히스토리로 계산만 더 하는 것이라 스캔 속도에 미치는 영향을 최소화했다.)
    """
    code = row['Code']
    name = row['Name']
    empty_tech_stats = [(tech_name, 0, 0, 0.0) for tech_name, _ in CONFIRMATION_TECHNIQUES]
    try:
        df = fdr.DataReader(code, start=start_date, end=end_date)
        if len(df) < 25:
            return None, empty_tech_stats

        ohlcv = df[['Open', 'High', 'Low', 'Close', 'Volume']].copy()

        # ---- 기법1/2/3에서 필요한 부가 정보 주입 (전체 시장 기준 1회 계산된 값) ----
        shares = row.get('Stocks', None)
        if pd.notna(shares) and shares:
            ohlcv['Shares'] = shares
        ohlcv['IsLeading'] = bool(row.get('IsLeading', False))
        ohlcv['MarketBullish'] = bool(market_bullish)

        today = df.iloc[-1]
        yesterday = df.iloc[-2]

        matched_names = []
        # ---- [신규] 이 종목에서 9개 기법 전체를 오늘 신호 + 과거 백테스트까지 한 번에 계산 ----
        # (오늘 신호가 뜬 종목만이 아니라 전 종목에 대해 계산해서 "기법별 전체시장 백테스트"에 사용)
        tech_stats = []
        matched_raw = {}
        for tech_name, check_fn in CONFIRMATION_TECHNIQUES:
            try:
                if check_fn(ohlcv):
                    matched_names.append(tech_name)
            except Exception as e:
                logging.warning(f"[{tech_name} 확인 오류] {name}({code}): {e}")

            try:
                wins, total, sum_returns = backtest_technique_raw(ohlcv, check_fn)
            except Exception as e:
                logging.warning(f"[{tech_name} 백테스트 오류] {name}({code}): {e}")
                wins, total, sum_returns = 0, 0, 0.0

            tech_stats.append((tech_name, wins, total, sum_returns))
            if tech_name in matched_names:
                matched_raw[tech_name] = (wins, total, sum_returns)

        if not matched_names:
            return None, tech_stats

        # ---- [신규] ATR 기반 손절가/목표가 ----
        stop_loss, target, rr_ratio = get_risk_levels(ohlcv)

        # ---- 오늘 매칭된 기법들의 (이 종목 자체) 과거 백테스트 결과를 합산해서 표시 ----
        if matched_raw:
            total_wins = sum(w for w, t, s in matched_raw.values())
            total_count = sum(t for w, t, s in matched_raw.values())
            total_sum_returns = sum(s for w, t, s in matched_raw.values())
            combined_win_rate, combined_avg_return = summarize_raw(total_wins, total_count, total_sum_returns)
            combined_sample = total_count
        else:
            combined_win_rate, combined_avg_return, combined_sample = None, None, 0

        change_rate = round(((today['Close'] - yesterday['Close']) / yesterday['Close']) * 100, 2)
        result = {
            '종목코드': code,
            '종목명': name,
            '시장': row['Market'],
            '종가': int(today['Close']),
            '등락률(%)': change_rate,
            '거래량': int(today['Volume']),
            '매수신호': len(matched_names),
            '충족조건': ", ".join(matched_names),
            '손절가': stop_loss,
            '목표가': target,
            '손익비': rr_ratio,
            '과거승률(%)': combined_win_rate,
            '평균수익률(%)': combined_avg_return,
            '과거신호횟수': combined_sample,
            '코스피시황': '상승' if market_bullish else '하락/조정',
        }
        return result, tech_stats
    except Exception:
        pass
    return None, empty_tech_stats


def fast_find_eaten_candles(max_workers=20):
    logging.info("전 종목 목록을 불러오는 중...")
    stocks = fdr.StockListing('KRX')
    stocks = stocks[stocks['Market'].isin(['KOSPI', 'KOSDAQ'])].copy()

    # ---- [신규] 부실/저유동성 종목 사전 제외 ----
    before_count = len(stocks)
    if 'Name' in stocks.columns:
        # 스팩(SPAC) 제외
        stocks = stocks[~stocks['Name'].str.contains('스팩', na=False)]
        # 우선주 제외 (이름이 '...우', '...우B', '...2우B' 등으로 끝나는 경우)
        stocks = stocks[~stocks['Name'].str.match(r'.*\d*우[A-Z]?$', na=False)]
    if 'Amount' in stocks.columns:
        stocks = stocks[stocks['Amount'] >= MIN_LIQUIDITY_AMOUNT]
    logging.info(f"부실/저유동성 종목 필터링: {before_count}개 -> {len(stocks)}개")

    end_date = datetime.today().strftime('%Y-%m-%d')
    # [변경] 일부 기법(예: year_high_lookback=240)이 과거 시점에서도 충분한 히스토리를
    # 확보할 수 있도록 조회 기간을 200일 -> 400일로 확장 (전체시장 백테스트 정확도 향상 목적)
    start_date = (datetime.today() - timedelta(days=400)).strftime('%Y-%m-%d')

    # ---- [기법2용] 거래대금(Amount) 상위 LEADING_STOCK_TOP_PCT 비율을 '주도주 후보군'으로 산정 ----
    # fdr.StockListing('KRX')가 이미 당일 거래대금을 담고 있어 추가 네트워크 호출이 필요 없다.
    if 'Amount' in stocks.columns:
        threshold = stocks['Amount'].quantile(1 - LEADING_STOCK_TOP_PCT)
        stocks['IsLeading'] = stocks['Amount'] >= threshold
    else:
        logging.warning("종목 목록에 'Amount'(거래대금) 컬럼이 없어 주도주 필터를 비활성화합니다.")
        stocks['IsLeading'] = False

    # ---- [기법3용] 코스피 시황(5일선) 1회 계산 ----
    market_bullish = get_market_bullish_flag(end_date)
    logging.info(f"코스피 시황(5일선 기준 상승 여부): {market_bullish}")

    results = []
    # [신규] 기법별 전체시장 집계용 누적 딕셔너리: {기법명: [wins합, total합, sum_returns합, 참여종목수]}
    tech_accum = {tech_name: [0, 0, 0.0, 0] for tech_name, _ in CONFIRMATION_TECHNIQUES}

    logging.info(f"총 {len(stocks)}개 종목을 {max_workers}개 스레드로 탐색합니다 (전체시장 기법 백테스트 포함)...")

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = [
            executor.submit(check_stock, row, start_date, end_date, market_bullish)
            for _, row in stocks.iterrows()
        ]
        for future in as_completed(futures):
            res, tech_stats = future.result()
            if res:
                results.append(res)
                logging.info(f"[1차 포착] {res['종목명']}({res['종목코드']}) | 등락률: {res['등락률(%)']}%")

            for tech_name, wins, total, sum_returns in tech_stats:
                acc = tech_accum[tech_name]
                acc[0] += wins
                acc[1] += total
                acc[2] += sum_returns
                if total > 0:
                    acc[3] += 1

    df_result = pd.DataFrame(results)
    # 과거승률이 높은 순으로 정렬(참고용) - 승률 데이터 없는 종목은 뒤로
    if not df_result.empty and '과거승률(%)' in df_result.columns:
        df_result = df_result.sort_values(
            by=['과거승률(%)', '매수신호'], ascending=[False, False], na_position='last'
        ).reset_index(drop=True)

    # ---- [신규] 기법별 전체시장 백테스트 요약 테이블 ----
    summary_rows = []
    for tech_name, (wins, total, sum_returns, stock_count) in tech_accum.items():
        win_rate, avg_return = summarize_raw(wins, total, sum_returns)
        summary_rows.append({
            '기법': tech_name,
            '신호횟수(전체시장)': total,
            '승률(%)': win_rate,
            '평균수익률(%)': avg_return,
            '신호발생종목수': stock_count,
            '검증기간(거래일)': MARKET_BACKTEST_LOOKBACK,
            '보유기간(거래일)': BACKTEST_FORWARD_DAYS,
        })
    df_technique_summary = pd.DataFrame(summary_rows).sort_values(
        by='승률(%)', ascending=False, na_position='last'
    ).reset_index(drop=True)

    return df_result, df_technique_summary


def save_excel(df: pd.DataFrame, df_technique_summary: pd.DataFrame = None) -> str:
    save_path = os.path.join(SCRIPT_DIR, f"양봉포착_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "포착종목"

    headers = [
        "종목명", "종목코드", "시장", "종가", "등락률(%)", "거래량", "매수신호", "충족조건",
        "손절가", "목표가", "손익비", "과거승률(%)", "평균수익률(%)", "과거신호횟수", "코스피시황",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    if not df.empty:
        for _, row in df.iterrows():
            ws.append([
                row["종목명"], row["종목코드"], row["시장"],
                row["종가"], row["등락률(%)"], row["거래량"], row["매수신호"], row["충족조건"],
                row.get("손절가"), row.get("목표가"), row.get("손익비"),
                row.get("과거승률(%)"), row.get("평균수익률(%)"), row.get("과거신호횟수"),
                row.get("코스피시황"),
            ])
            r = ws.max_row

            name_cell = ws.cell(row=r, column=1)
            name_cell.hyperlink = f"https://finance.naver.com/item/main.naver?code={row['종목코드']}"
            name_cell.font = Font(color="0563C1", underline="single")

            code_cell = ws.cell(row=r, column=2)
            youtube_url = f"https://www.youtube.com/results?search_query={quote(row['종목명'] + ' 주가')}"
            code_cell.hyperlink = youtube_url
            code_cell.font = Font(color="FF0000", underline="single")

    for col_cells in ws.columns:
        length = max(len(str(c.value)) for c in col_cells if c.value is not None)
        ws.column_dimensions[col_cells[0].column_letter].width = max(10, length + 4)

    # ---- [신규] 2번째 시트: 기법별 전체시장 백테스트 요약 ----
    if df_technique_summary is not None and not df_technique_summary.empty:
        ws2 = wb.create_sheet("기법별_전체시장백테스트")
        summary_headers = list(df_technique_summary.columns)
        ws2.append(summary_headers)
        for cell in ws2[1]:
            cell.font = Font(bold=True)
        for _, row in df_technique_summary.iterrows():
            ws2.append([row[col] for col in summary_headers])
        for col_cells in ws2.columns:
            length = max(len(str(c.value)) for c in col_cells if c.value is not None)
            ws2.column_dimensions[col_cells[0].column_letter].width = max(10, length + 4)

    wb.save(save_path)
    return save_path


def send_mail(excel_path: str, stock_count: int):
    today_str = datetime.now().strftime('%Y-%m-%d')
    msg = EmailMessage()
    msg["Subject"] = f"[양봉 포착 결과] {today_str} - 총 {stock_count}건"
    msg["From"] = NAVER_EMAIL
    msg["To"] = RECEIVER_EMAIL
    msg.set_content(
        f"{today_str} 스캔 결과입니다.\n"
        f"총 {stock_count}개 종목이 포착되었습니다.\n"
        f"첨부된 엑셀 파일을 확인해주세요.\n"
        f"(종목명 클릭 시 네이버 차트로, 종목코드 클릭 시 유튜브 검색으로 이동합니다)\n"
        f"※ '손절가/목표가'는 ATR 기반 참고치, '과거승률/평균수익률'은 해당 종목 자체의\n"
        f"  과거 신호 재현성 참고 지표입니다(표본이 적으면 신뢰도가 낮음).\n"
        f"※ 2번째 시트 '기법별_전체시장백테스트'에는 오늘 신호와 무관하게 스캔한 전 종목을\n"
        f"  대상으로 각 기법이 최근 {MARKET_BACKTEST_LOOKBACK}거래일 동안 실제로 얼마나 잘 맞았는지\n"
        f"  집계한 결과가 있습니다. 승률이 지속적으로 낮은 기법은 제외를 고려해보세요.\n"
        f"  실제 투자 판단과 리스크 관리는 반드시 본인 책임 하에 하시기 바랍니다."
    )

    with open(excel_path, "rb") as f:
        msg.add_attachment(
            f.read(),
            maintype="application",
            subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=os.path.basename(excel_path),
        )

    context = ssl.create_default_context()
    with smtplib.SMTP("smtp.naver.com", 587) as server:
        server.starttls(context=context)
        server.login(NAVER_EMAIL, NAVER_PASSWORD)
        server.send_message(msg)


def main():
    start_time = datetime.now()
    logging.info("===== 자동 실행 시작 =====")
    try:
        df_result, df_technique_summary = fast_find_eaten_candles(max_workers=20)
        excel_path = save_excel(df_result, df_technique_summary)
        logging.info(f"엑셀 저장 완료: {excel_path}")

        send_mail(excel_path, len(df_result))
        logging.info("메일 발송 완료")
    except Exception as e:
        logging.exception(f"실행 중 오류 발생: {e}")
    finally:
        logging.info(f"소요 시간: {datetime.now() - start_time}")
        logging.info("===== 자동 실행 종료 =====\n")


if __name__ == "__main__":
    main()
